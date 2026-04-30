#!/usr/bin/env python3
"""
Module 0 – Data Pipeline Foundation (Mask-Aware)
==================================================
Parses 14 monthly OSOS OSF Excel files, extracts Cekis (consumption) data,
reshapes to long format, adds per-hour is_missing flag, cross-references
BEDAS outage reports, and produces a data quality summary.

Outputs:
  - master_dataset.csv       : ML-ready long-format dataset with is_missing flag
  - outage_events.csv        : Parsed BEDAS modem+meter outage events
  - missing_data_summary.csv : Per-day missing-data metrics and data_quality_flag
  - data_quality_report.csv  : Summary-level quality metrics
"""

import glob
import os
import re
import sys
from datetime import date, datetime, timedelta
from calendar import monthrange

import openpyxl
import numpy as np
import pandas as pd

# ---------------------------------------------------------------------------
# Config  (DATA_DIR is set by unified_preprocessing.py before import)
# ---------------------------------------------------------------------------
DATA_DIR = os.environ.get("BEDAS_DATA_DIR",
                          os.path.dirname(os.path.abspath(__file__)))
FILE_PATTERN = os.path.join(DATA_DIR, "osf_*.xlsx")

# Excel layout constants
DATA_ROW_START = 23
DATA_ROW_END   = 70
DAY_COL_START  = 3
DAY_COL_END    = 33
HOUR_COL       = 1
TYPE_COL       = 2

NIGHT_HOURS = [22, 23, 0, 1, 2, 3, 4, 5]
DAY_HOURS   = [9, 10, 11, 12, 13, 14, 15, 16]

# Valid data period ends on this date (inclusive)
CUTOFF = pd.Timestamp("2026-02-13")

# BEDAS report files
MODEM_OUTAGE_FILES = [
    "rpt-301_modem_kesinti_raporu_(kesinti_bazli)_f2Umn.xlsx",
    "rpt-308_modem_kesinti_raporu_(butun_kesintiler)_bPZj6.xlsx",
]
METER_OUTAGE_FILE = "rpt-300_sayac_kesinti_raporu_(kesinti_bazli)_0XldY.xlsx"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def parse_filename(fname: str):
    """Extract OSF_ID, month, year from filename like osf_2193681000_01_2025_xxx.xlsx"""
    base = os.path.basename(fname)
    m = re.match(r"osf_(\d+)_(\d{2})_(\d{4})_", base)
    if not m:
        return None, None, None
    return m.group(1), int(m.group(2)), int(m.group(3))


def parse_hour(hour_str: str) -> int:
    """Convert '00-01' -> 0, '13-14' -> 13, etc."""
    if hour_str is None:
        return None
    hour_str = str(hour_str).strip()
    m = re.match(r"(\d{1,2})\s*[-\u2013]\s*\d{1,2}", hour_str)
    if m:
        return int(m.group(1))
    return None


def safe_float(val):
    """Convert cell value to float. Returns NaN for None/empty (true missing data)."""
    if val is None:
        return float("nan")
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if s == "" or s.startswith("="):
        return float("nan")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return float("nan")


def parse_turkish_datetime(dt_str: str) -> datetime:
    """Parse '09/01/2025 09:01:39' (DD/MM/YYYY HH:MM:SS) -> datetime."""
    dt_str = str(dt_str).strip()
    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"]:
        try:
            return datetime.strptime(dt_str, fmt)
        except (ValueError, TypeError):
            continue
    return None


# ---------------------------------------------------------------------------
# BEDAS Outage Report Parsing
# ---------------------------------------------------------------------------
def parse_outage_reports() -> pd.DataFrame:
    """Parse BEDAS modem and meter outage reports into a unified event list."""
    events = []

    for fname in MODEM_OUTAGE_FILES:
        fpath = os.path.join(DATA_DIR, fname)
        if not os.path.exists(fpath):
            continue
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            osf_id = ws.cell(row=row, column=1).value
            start_str = ws.cell(row=row, column=9).value
            end_str = ws.cell(row=row, column=10).value
            duration_sec = ws.cell(row=row, column=11).value
            if osf_id is None or start_str is None:
                continue
            start_dt = parse_turkish_datetime(str(start_str))
            end_dt = parse_turkish_datetime(str(end_str))
            if start_dt is None:
                continue
            events.append({
                "OSF_ID": str(osf_id),
                "Source": "modem_outage",
                "Start": start_dt,
                "End": end_dt,
                "Duration_sec": int(duration_sec) if duration_sec else None,
                "SourceFile": fname,
            })
        wb.close()

    fpath = os.path.join(DATA_DIR, METER_OUTAGE_FILE)
    if os.path.exists(fpath):
        wb = openpyxl.load_workbook(fpath, data_only=True)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            osf_id = ws.cell(row=row, column=1).value
            start_str = ws.cell(row=row, column=10).value
            end_str = ws.cell(row=row, column=11).value
            duration_min = ws.cell(row=row, column=12).value
            if osf_id is None or start_str is None:
                continue
            start_dt = parse_turkish_datetime(str(start_str))
            end_dt = parse_turkish_datetime(str(end_str))
            if start_dt is None:
                continue
            events.append({
                "OSF_ID": str(osf_id),
                "Source": "meter_outage",
                "Start": start_dt,
                "End": end_dt,
                "Duration_sec": int(duration_min) * 60 if duration_min else None,
                "SourceFile": METER_OUTAGE_FILE,
            })
        wb.close()

    if not events:
        return pd.DataFrame(columns=["OSF_ID", "Source", "Start", "End",
                                      "Duration_sec", "SourceFile"])
    df = pd.DataFrame(events)
    df = df.drop_duplicates(subset=["OSF_ID", "Start", "End"]).sort_values("Start")
    df.reset_index(drop=True, inplace=True)
    return df


def check_outage_overlap(dt: date, hour: int, outage_df: pd.DataFrame, osf_id: str) -> bool:
    """Check if a given (date, hour) falls within any outage event."""
    if outage_df.empty:
        return False
    hour_start = datetime(dt.year, dt.month, dt.day, hour, 0, 0)
    hour_end = hour_start + timedelta(hours=1)
    osf_outages = outage_df[outage_df["OSF_ID"] == osf_id]
    for _, evt in osf_outages.iterrows():
        if evt["Start"] is not None and evt["End"] is not None:
            if evt["Start"] < hour_end and evt["End"] > hour_start:
                return True
    return False


# ---------------------------------------------------------------------------
# Core pipeline
# ---------------------------------------------------------------------------
def extract_file(filepath: str) -> pd.DataFrame:
    """Parse one OSF Excel file and return a long-format DataFrame."""
    osf_id, month, year = parse_filename(filepath)
    if osf_id is None:
        print(f"  WARNING: Skipping (cannot parse filename): {filepath}")
        return pd.DataFrame()

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    days_in_month = monthrange(year, month)[1]

    records = []

    for row_idx in range(DATA_ROW_START, DATA_ROW_END + 1):
        cell_type = ws.cell(row=row_idx, column=TYPE_COL).value
        if cell_type is None or str(cell_type).strip().lower() != "\u00e7eki\u015f":
            continue

        hour_val = ws.cell(row=row_idx, column=HOUR_COL).value
        if hour_val is None and row_idx > DATA_ROW_START:
            hour_val = ws.cell(row=row_idx - 1, column=HOUR_COL).value
        hour = parse_hour(hour_val)
        if hour is None:
            continue

        for day in range(1, days_in_month + 1):
            col = DAY_COL_START + day - 1
            raw = ws.cell(row=row_idx, column=col).value
            consumption = safe_float(raw)
            is_missing = 1 if np.isnan(consumption) else 0
            dt = date(year, month, day)
            records.append({
                "Date": dt,
                "Hour": hour,
                "Consumption_kWh": consumption,
                "is_missing": is_missing,
                "OSF_ID": osf_id,
                "Month": month,
                "Year": year,
            })

    wb.close()
    return pd.DataFrame(records)


def add_missingness_cause(master: pd.DataFrame, outage_df: pd.DataFrame) -> pd.DataFrame:
    """Label each missing hour as outage-related or unknown."""
    causes = []
    missing_mask = master["is_missing"] == 1
    for idx in master[missing_mask].index:
        row = master.loc[idx]
        dt = row["Date"]
        if isinstance(dt, pd.Timestamp):
            dt = dt.date()
        hour = int(row["Hour"])
        osf_id = str(row["OSF_ID"])
        if check_outage_overlap(dt, hour, outage_df, osf_id):
            causes.append("measurement_missing_due_to_outage")
        else:
            causes.append("unknown_missingness")

    master["missingness_cause"] = ""
    master.loc[missing_mask, "missingness_cause"] = causes
    return master


def build_missing_data_summary(master: pd.DataFrame) -> pd.DataFrame:
    """Per-day missing data metrics and quality flag."""
    rows = []
    for (osf, dt), grp in master.groupby(["OSF_ID", "Date"]):
        n_total = len(grp)
        n_missing = int(grp["is_missing"].sum())
        missing_ratio = round(n_missing / 24, 4)
        missing_hours = sorted(grp.loc[grp["is_missing"] == 1, "Hour"].tolist())
        outage_count = int((grp["missingness_cause"] == "measurement_missing_due_to_outage").sum())
        unknown_count = int((grp["missingness_cause"] == "unknown_missingness").sum())

        if missing_ratio == 0:
            flag = "complete"
        elif missing_ratio <= 0.25:
            flag = "partial_missing"
        else:
            flag = "data_quality_low"

        rows.append({
            "OSF_ID": osf,
            "Date": dt,
            "n_present": 24 - n_missing,
            "n_missing": n_missing,
            "missing_ratio": missing_ratio,
            "missing_hours": missing_hours,
            "outage_overlap_count": outage_count,
            "unknown_missing_count": unknown_count,
            "data_quality_flag": flag,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Main (callable from unified_preprocessing or standalone)
# ---------------------------------------------------------------------------
def run_pipeline(data_dir: str = None, output_dir: str = None):
    """
    Run the full data pipeline.

    Parameters
    ----------
    data_dir   : directory containing osf_*.xlsx and rpt-*.xlsx files
    output_dir : directory to write output CSVs (defaults to data_dir)

    Returns
    -------
    master, outage_df, missing_summary : DataFrames
    """
    global DATA_DIR, FILE_PATTERN
    if data_dir:
        DATA_DIR = data_dir
        FILE_PATTERN = os.path.join(DATA_DIR, "osf_*.xlsx")
    if output_dir is None:
        output_dir = DATA_DIR

    print("=" * 70)
    print(" Module 0 - Data Pipeline Foundation (Mask-Aware)")
    print("=" * 70)

    # 1. Parse BEDAS outage reports
    print("\nParsing BEDAS outage reports...")
    outage_df = parse_outage_reports()
    print(f"   Found {len(outage_df)} outage events")
    outage_df.to_csv(os.path.join(output_dir, "outage_events.csv"), index=False)

    # 2. Discover OSF files
    files = sorted(glob.glob(FILE_PATTERN))
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    print(f"\nFound {len(files)} OSF Excel files\n")

    if not files:
        print("ERROR: No files found. Check the directory.")
        sys.exit(1)

    # 3. Parse each file
    all_dfs = []
    for fp in files:
        osf_id, month, year = parse_filename(fp)
        label = f"{year}-{month:02d}" if year else "?"
        print(f"  Processing {os.path.basename(fp)}  ({label})")
        df = extract_file(fp)
        if not df.empty:
            all_dfs.append(df)
            n_missing = int(df["is_missing"].sum())
            print(f"     -> {len(df)} records ({n_missing} missing)")
        else:
            print(f"     -> WARNING: No records extracted")

    if not all_dfs:
        print("\nERROR: No data extracted from any file.")
        sys.exit(1)

    # 4. Combine
    master = pd.concat(all_dfs, ignore_index=True)
    master["Date"] = pd.to_datetime(master["Date"])
    master.sort_values(["OSF_ID", "Date", "Hour"], inplace=True)
    master.reset_index(drop=True, inplace=True)

    # Apply cutoff
    before_cutoff = len(master)
    master = master[master["Date"] <= CUTOFF].reset_index(drop=True)
    removed_days = (before_cutoff - len(master)) // 24
    if removed_days > 0:
        print(f"   Removed {removed_days} days after cutoff {CUTOFF.date()}")

    print(f"\nCombined master dataset: {len(master)} rows")
    print(f"   Date range: {master['Date'].min().date()} -> {master['Date'].max().date()}")
    print(f"   Total missing hours: {int(master['is_missing'].sum())}")

    # 5. Add missingness cause
    print("\nCross-referencing missing hours with outage events...")
    master = add_missingness_cause(master, outage_df)
    outage_related = (master["missingness_cause"] == "measurement_missing_due_to_outage").sum()
    unknown_related = (master["missingness_cause"] == "unknown_missingness").sum()
    print(f"   Outage-related missing: {outage_related}")
    print(f"   Unknown missing: {unknown_related}")

    # 6. Missing data summary
    print("\nBuilding missing data summary...")
    missing_summary = build_missing_data_summary(master)

    for flag in ["complete", "partial_missing", "data_quality_low"]:
        n = (missing_summary["data_quality_flag"] == flag).sum()
        print(f"   {flag:25s}: {n} days")

    # 7. Export
    master.to_csv(os.path.join(output_dir, "master_dataset.csv"), index=False)
    missing_summary.to_csv(os.path.join(output_dir, "missing_data_summary.csv"), index=False)

    total_days = len(missing_summary)
    complete_days = int((missing_summary["data_quality_flag"] == "complete").sum())
    partial_days = int((missing_summary["data_quality_flag"] == "partial_missing").sum())
    low_quality_days = int((missing_summary["data_quality_flag"] == "data_quality_low").sum())

    summary = pd.DataFrame({
        "Metric": [
            "Total_Days", "Complete_Days", "Partial_Missing_Days",
            "Data_Quality_Low_Days", "Total_Missing_Hours",
            "Outage_Related_Missing", "Unknown_Missing",
            "Completeness_Pct",
        ],
        "Value": [
            total_days, complete_days, partial_days,
            low_quality_days, int(master["is_missing"].sum()),
            int(outage_related), int(unknown_related),
            round(complete_days / total_days * 100, 2) if total_days else 0,
        ],
    })
    summary.to_csv(os.path.join(output_dir, "data_quality_report.csv"), index=False)

    print(f"\nFiles saved to: {output_dir}")
    print("=" * 70)

    return master, outage_df, missing_summary


def main():
    run_pipeline()


if __name__ == "__main__":
    main()
